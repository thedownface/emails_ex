
from fileinput import filename
import scrapy
from scrapy.spiders import CrawlSpider, Request
import re
from scrapy_selenium import SeleniumRequest
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from time import sleep
import xlsxwriter
from bs4 import BeautifulSoup
from scrapy.crawler import CrawlerProcess
import os
from selenium import webdriver
from pathlib import Path
from tldextract import extract
from scrapy.crawler import CrawlerProcess
from django.views import View
from django.shortcuts import HttpResponse,render


class EmailExtractor(CrawlSpider):
    name='scrapy_run'
    def __init__(self):
        self.queries=[]
        self.emaillist=[]
        self.row=0
        self.save_file="emails.txt"
    
    def read_excel(self,file):
        wb=openpyxl.load_workbook(file)
      
        sh=wb.active
        for i in range(1,sh.max_row+1):
            cell_obj=sh.cell(row=i,column=1)
            tsd, td, tsu = extract(cell_obj.value)
            search_query=td + '.' + tsu
            #pass_val='"@'+str(search_query)+'" Email Address'
            self.queries.append(search_query)
            return self.queries

    def start_requests(self):
        

        WINDOW_SIZE="1920,1080"
        
        path="C:/Users/iamfa/OneDrive/Desktop/SCRAPY/email_extraction/email_extraction/spiders/msedgedriver.exe"
        options=webdriver.EdgeOptions()
        #options.add_argument("--headless")
        #options.add_argument("--window-size=%s" % WINDOW_SIZE)
        options.add_argument('--ignore-ssl-errors=yes')
        options.add_argument('--ignore-certificate-errors')
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        browser=webdriver.Edge(options=options,executable_path=path)
        url_list=[]
        for query in self.read_excel():  # iterate through queries
            url="https://www.bing.com/search?q=%40"+str(query)+"+%22Email+Address%22"
            try:
                browser.get(url)
                links=browser.find_elements(By.TAG_NAME,'cite')    
                for link in links:
                    url_list.append(link.text)
            except:
                continue
                

            resultno=0
            for results in url_list:
                
                if resultno==10:
                    break
                try:
                    resultno+=1
                    yield SeleniumRequest(
                        url=results,
                        callback=self.parse,
                        wait_until=EC.presence_of_element_located(
                            (By.TAG_NAME, "html")),
                        dont_filter=True
                    )
                except:
                    continue
            url_list.clear()
   
    
    def parse(self, response):
        
        file1=open(self.save_file,'a')
        
        EMAIL_REGEX =r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+"
        emails = re.finditer(EMAIL_REGEX, str(response.text))
        for email in emails:
            self.emaillist.append(email.group())
            
        for email in set(self.emaillist):
            if "png" or ".svg" or ".webp" or ".jpg" or ".jpeg" or ".wixpress" not in email:
                if "j" and "doe" not in email:
                    file1.write(email+"\n")
                #yield{
                    #"emails": email
                #}
            
        self.emaillist.clear()