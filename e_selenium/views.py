
# Create your views here.
from doctest import OutputChecker
import io
from urllib import response
from django.shortcuts import HttpResponse, render
import re#for finding emails from text
import xlsxwriter as xs#for writing in an excel sheet
from bs4 import BeautifulSoup#for scraping html data
import openpyxl#for reading excel
from urllib.parse import urlparse#for parsing website url and finding domain name
from selenium import webdriver#for calling browser
from pathlib import Path
from tldextract import extract


def e_selenium(request):
    response = HttpResponse(content_type = 'applications/ms-excel')
    response['Content-Disposition'] = 'attachment; filename = "emails.xlsx"'
    if "GET" == request.method:
        return render(request, 'e_selenium/selenium.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        WINDOW_SIZE="1920,1080"

        path='driver/msedgedriver.exe'
        options=webdriver.EdgeOptions()
        #options.add_argument("--headless")
        #options.add_argument("--window-size=%s" % WINDOW_SIZE)
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        #It is used to append an experimental option which is passed to the Chrome browser.
        browser=webdriver.Edge(options=options,executable_path=path)

            
        # To open the Workbook 
        #Reading an excel sheet containing website links    
        wb = openpyxl.load_workbook(excel_file)     
        sh=wb.active

        #export_file=Path(excel_file).stem+"_emails.xlsx"

        #Creating an excel sheet to save emaid ids
        output = io.BytesIO()
        workbook=xs.Workbook(output)
        sheet=workbook.add_worksheet("first")
        excel_data=[]   
        # For row 0 and column 0     
        row=0
        for i in range(1,sh.max_row+1):
            cell_obj=sh.cell(row=i,column=1)
            
        
            try:
                tsd, td, tsu = extract(cell_obj.value)
                search_query=td + '.' + tsu
            except:
                continue


            
            
            url='https://www.bing.com/search?q=%40'+str(search_query)+'+%22Email+Address%22"'

            browser.get(url)
            html=browser.page_source
            soup=BeautifulSoup(html,'lxml')
            emails = re.findall(r"[a-z0-9\.\-+]+@[a-z0-9\.\-+]+\.[a-z]+", str(soup.text))
           

        
            if len(emails)==0:
                sheet.write(row,2,search_query)
                
            for email in emails:
                if "doe" not in email:
                    sheet.write(row,1,email)
                    row+=1
                    excel_data.append(email)

        browser.close()
        workbook.close()
        output.seek(0)
        save_file = str(excel_file).split('.')
        file_name=save_file[0]+".xlsx"
        response = HttpResponse(output,content_type ='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % file_name
        
        return response
