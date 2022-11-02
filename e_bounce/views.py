from django.shortcuts import render
import email
from email.header import decode_header
import imaplib
import io
import openpyxl
import xlsxwriter as xs
from django.shortcuts import HttpResponse


def e_bounce(request): 
    
    if request.method=='POST':
        show_name=request.POST.get('subject')
        row=0
        server ="imap.gmail.com"                     
        imap = imaplib.IMAP4_SSL(server)
        output = io.BytesIO()
        workbook=xs.Workbook(output)
        sheet=workbook.add_worksheet("first")
        excel_data=[]  
        # intantiate the username and the password
        wb = openpyxl.load_workbook("DOCUMENTS\email ids.xlsx")
        sh=wb.active
        for i in range(1,8):
            flag=0
            if(i==7):
                i=1
                continue
            sender=(sh.cell(row=i,column=1))
            sender_password= (sh.cell(row=i,column=2))
            username=str(sender.value)
            password=str(sender_password.value)
        
        # login into the gmail account
            imap.login(username, password)               
        
            # select the e-mails
            res, messages = imap.select('"[Gmail]/Sent Mail"')   
            
            # calculates the total number of sent messages
            messages = int(messages[0])
            
            # determine the number of e-mails to be fetched
            n=100
            
            # iterating over the e-mails
            for i in range(messages, messages - n, -1):
                res, msg = imap.fetch(str(i), "(RFC822)")  
                if flag==1:
                    break   
                for response in msg:
                    if isinstance(response, tuple):
                        msg = email.message_from_bytes(response[1])
                        
                        # getting the sender's mail id
                        From = msg["From"]
            
                        # getting the subject of the sent mail
                        subject = msg["Subject"]
                        if show_name in subject:
                            row+=1
                            sheet.write(row,1,str(msg["To"]))
                        else:
                            flag=1
                            break
                  
                        
                            

                        
        workbook.close()
        output.seek(0)
        file_name = "email.xlsx"
        response = HttpResponse(output,content_type ='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
            
        response['Content-Disposition'] = 'attachment; filename=%s' % file_name
        return response
    else:
        return render(request, 'e_bounce/bounce.html')
