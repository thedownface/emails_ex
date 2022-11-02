from tkinter import E
from django.shortcuts import render,HttpResponse
import smtplib
import openpyxl
from termcolor import colored
from email.message import EmailMessage
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart



def e_sender(request):

    if request.method == 'POST':
        excel_file = request.FILES["excel_file"]
        subject=request.POST.get("subject")
        attendees=request.POST.get("attendees")
        showname=request.POST.get("showname")
        sent_number=1
        wb = openpyxl.load_workbook("DOCUMENTS\email ids.xlsx")
        sh=wb.active
        sent=0
        MAX_EMAILS=1900
        count=0
        email_list=openpyxl.load_workbook(excel_file)
        email_sheet=email_list.active
        for i in range(1,8):
            if(i==7):
                i=1
                continue
            sender=(sh.cell(row=i,column=1))
            sender_password= (sh.cell(row=i,column=2))

            #LOGIN TRY
            
            server=smtplib.SMTP_SSL('smtp.gmail.com',465)
            server.ehlo()
            try:
                server.login(str(sender.value),str(sender_password.value))
            except:
                continue
            """except:
                print('wrong emailid or password')
                print('try again')
                break"""

            

        

            
            #reading a message


            sender_name=sh.cell(row=i,column=3)
            sender_profession=sh.cell(row=i,column=4)
            

            html="""
                <html>
                <body>
                <p>Hello, </p>
                <p>Hope you are doing well.</p>
                <p>I see that your company is a part of <b>"""+showname+ """</b>.</p>
                <div>So I’m wondering if you are interested in acquiring Attendees List contacts 
                of<b> """ +attendees +"""</b> for your pre and post event campaigns.</div>
                <div>We provide all the relevant information about the attendee including
                Company Name, Company URL, Contact Name, Title, Phone number, Fax 
                Number, Email Address, Company Address and Industry type. Hence you 
                can use this information for your email, telephonic and mailing campaigns.</div>
                <p></p>
                <div> If you are interested, drop me a line. We will get back to you with pricing
                counts and other information for your review.</div>
                <p>If you think I should be talking to someone else, please forward this email </p>
                <p>to the concerned person.</p>
                <p>Looking forward to hearing from you.</p>
                <p>Best Regards,<p>
                <b><p style="color:hsl(270, 100%, 50%);">"""+sender_name.value+"""</p></b>
                <b><p style="color:hsl(270, 100%, 50%);">"""+sender_profession.value+"""</p></b>
                <p style="color:rgb(220,220,220);font-size:12px;">If you don't wish to receive emails from us reply back with <span style="hsl(278, 100%, 50%);"> Unsubscribe</span></p>
                </body></html>"""


            
        

            #reading subject
            
            #body="Subject:{}\n\n{}".format(subject,msg)
            if(sent==500 or MAX_EMAILS<=0):
                print("id",i ,"have ",MAX_EMAILS-500,"emails left")
                MAX_EMAILS=MAX_EMAILS-500
                sent=0
                continue
            
            for j in range(sent_number,email_sheet.max_row+1):
                count=count+1 
                email=(email_sheet.cell(row=j,column=2))
                if(count==sh.max_row+1):
                    break

                emailmsg = MIMEMultipart("alternative")
                emailmsg['Subject']=str(subject).upper()
                emailmsg['From']=str(sender.value)
                emailmsg['To']=str(email.value)
                part = MIMEText(html, "html")
                emailmsg.attach(part)
                try:
                    server.send_message(emailmsg)
                    if smtplib.SMTPRecipientsRefused:
                        print("REFUSED")
                        print(email.value)
                except:
                    sent_number=sent
                    break
                
                        
                
            print("invalid emails")  
            print('total number of valid emails sent=',+count)
        return HttpResponse('SENT'+str(sent)+'EMAILS')
    else:
        return render(request, 'e_sender/sender.html')